from flask import Flask, render_template, request, send_file, flash, redirect, url_for
import os
import io
import json
import re
import requests
from openpyxl import load_workbook
from openpyxl.utils import range_boundaries
import gspread
from oauth2client.service_account import ServiceAccountCredentials

app = Flask(__name__)

# --- Flask Secret Key 読み込み ---
with open("/etc/secrets/flask_secret_key", "r") as f:
    app.secret_key = f.read().strip()

# --- Google認証（Secret File 経由） ---
CREDENTIAL_FILE_PATH = "/etc/secrets/credentials.json"

def get_credentials():
    with open(CREDENTIAL_FILE_PATH, "r", encoding="utf-8") as f:
        credentials_dict = json.load(f)
    scope = ["https://spreadsheets.google.com/feeds", "https://www.googleapis.com/auth/drive"]
    return ServiceAccountCredentials.from_json_keyfile_dict(credentials_dict, scope)

# --- テキスト抽出 ---
def extract_data(text):
    patterns = {
        "製造番号": r"製造番号[:：]\s*([^\s)]+)",
        "印刷番号": r"印刷番号[:：]\s*([^\s\n]+)",
        "製造日": r"製造日[:：]\s*(.+?)(?:\n|$)",
        "会社名": r"会社名[:：]\s*(.+?)(?:\n|$)",
        "製品名": r"製品名[:：]\s*(.+?)(?:\n|$)",
        "製品種類": r"製品種類[:：]\s*(.+?)(?:\n|$)",
        "外装包材": r"外装包材[:：]\s*(.+?)(?:\n|$)",
        "表面印刷": r"表面印刷[:：][^\n]+.*?表面印刷[:：]\s*(.+?)(?:\n|$)",
        "製造個数": r"製造個数[:：]\s*(.+?)(?:\n|$)",
        "ファイル名": r"<印刷用データ\(\.FMT\)>.*?ファイル名[:：]\s*(.+?)(?:\n|$)",
        "印刷データ（元）": r"印刷データ[:：]\s*(.+?)(?:\n|$)"
    }

    results = {}
    for key, pattern in patterns.items():
        match = re.search(pattern, text, re.DOTALL)
        if match:
            results[key] = match.group(1).strip()

    if "印刷データ（元）" in results:
        raw = results.pop("印刷データ（元）")
        results["印刷データ"] = "リピート" if "従来の" in raw else "新規"
    else:
        results["印刷データ"] = ""

    return results

# --- メインエンドポイント ---
@app.route("/", methods=["GET", "POST"])
def index():
    extracted_data = {}
    if request.method == "POST":
        text = request.form["text"]
        extracted_data = extract_data(text)

        template_path = "printlist_form.xlsx"
        wb = load_workbook(template_path)
        ws = wb.active

        cell_map = {
            "製造日": "B2",
            "製造番号": "E1",
            "印刷番号": "E2",
            "会社名": "B3",
            "製品名": "B4"
        }

        for key, cell in cell_map.items():
            if key in extracted_data:
                target_cell = ws[cell]
                if hasattr(target_cell, "merged_cells") or type(target_cell).__name__ == "MergedCell":
                    for merged_range in ws.merged_cells.ranges:
                        if cell in merged_range:
                            min_col, min_row, _, _ = range_boundaries(str(merged_range))
                            top_left_cell = ws.cell(row=min_row, column=min_col)
                            top_left_cell.value = extracted_data[key]
                            break
                else:
                    ws[cell] = extracted_data[key]

        excel_stream = io.BytesIO()
        wb.save(excel_stream)
        excel_stream.seek(0)

        creds = get_credentials()
        client = gspread.authorize(creds)

        SPREADSHEET_ID = "1fKN1EDZTYOlU4OvImQZuifr2owM8MIGgQIr0tu_rX0E"
        ss = client.open_by_key(SPREADSHEET_ID)
        template_ws = ss.worksheet("sheet1")
        output_ws = ss.worksheet("printlist")

        existing_rows = len(output_ws.get_all_values())
        block_index = max((existing_rows - 2) // 10, 0)
        start_row = block_index * 10 + 1

        template_range = template_ws.get_values("A1:N10")
        for i, row in enumerate(template_range):
            output_ws.update(f"A{start_row + i}:N{start_row + i}", [row])

        sheet_map = {
            "A3": "印刷データ",
            "B3": "ファイル名",
            "C3": "製造番号",
            "C7": "印刷番号",
            "D3": "製造日",
            "E3": "会社名",
            "E5": "製品名",
            "G3": "製品種類",
            "G6": "外装包材",
            "G9": "表面印刷",
            "L3": "製造個数"
        }

        for cell_a1, key in sheet_map.items():
            if key in extracted_data:
                row = int(cell_a1[1:])
                col = ord(cell_a1[0].upper()) - 65 + 1
                output_ws.update_cell(start_row + (row - 1), col, extracted_data[key])

        return send_file(
            excel_stream,
            as_attachment=True,
            download_name="output.xlsx",
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    return render_template("index.html")

# --- Google Apps Script 側の clearData を呼び出すエンドポイント ---
@app.route("/clear", methods=["POST"])
def clear_sheet():
    GAS_ENDPOINT = "https://script.google.com/macros/s/AKfycbyhVDrk1fweJSj3UkoXL9m1tHIRcK4iMIo_IQwJcNN7phZNGg5513NtuQy-ROf7Qig4/exec"
    try:
        response = requests.post(GAS_ENDPOINT)
        if response.status_code == 200 and response.text.strip() == "CLEARED":
            flash("スプレッドシートのデータをクリアしました。")
        else:
            flash("クリアに失敗しました：" + response.text)
    except Exception as e:
        flash("通信エラー：" + str(e))
    return redirect(url_for("index"))

# --- Google Apps Script 側のテンプレートブロックコピーを呼び出すエンドポイント ---
@app.route("/copy", methods=["POST"])
def copy_template_block():
    GAS_ENDPOINT = "https://script.google.com/macros/s/AKfycbxcr6gSE06BXBOMZnuRXpPYEJk1VC-Ei7qSR5jDNoLCFnDR2tXXzvzLTKDi0iyLpqgo/exec"
    try:
        response = requests.post(GAS_ENDPOINT)
        if response.status_code == 200 and "TEMPLATE COPIED" in response.text:
            flash("テンプレートをコピーしました。")
        else:
            flash("コピーに失敗しました：" + response.text)
    except Exception as e:
        flash("通信エラー：" + str(e))
    return redirect(url_for("index"))

if __name__ == "__main__":
    app.run(debug=True)
