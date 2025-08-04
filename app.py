from flask import Flask, render_template, request, send_file, flash, redirect, url_for
import os
import io
import json
import re
import requests
import logging
# ✅ ログの基本設定
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s"
)
from openpyxl import load_workbook
from openpyxl.utils import range_boundaries
import gspread
from oauth2client.service_account import ServiceAccountCredentials
from style_writer import apply_template_style  # 追加

app = Flask(__name__)

# --- グローバル定数 ---
GAS_ENDPOINT = "https://script.google.com/macros/s/AKfycbyl1YhHikED4ogJ_z4iC7MDRXRevn9KeMEpbwfgxSDq9iEKSOpUSkrxtrupXtECFHAm/exec"
SPREADSHEET_ID = "1fKN1EDZTYOlU4OvImQZuifr2owM8MIGgQIr0tu_rX0E"
TEMPLATE_ROW_HEIGHT = 8        # テンプレート1ブロックの行数（A1:O8）
TEMPLATE_DATA_ROW_OFFSET = 3   # テンプレート内でデータが始まる行（3行目）
TEMPLATE_EXCEL_PATH = "printlist_form.xlsx"

# --- Flask Secret Key 読み込み ---
with open("/etc/secrets/flask_secret_key", "r") as f:
    app.secret_key = f.read().strip()

# --- Google認証 ---
CREDENTIAL_FILE_PATH = "/etc/secrets/credentials.json"

def get_credentials():
    with open(CREDENTIAL_FILE_PATH, "r", encoding="utf-8") as f:
        credentials_dict = json.load(f)
    scope = ["https://spreadsheets.google.com/feeds", "https://www.googleapis.com/auth/drive"]
    return ServiceAccountCredentials.from_json_keyfile_dict(credentials_dict, scope)

# --- テキスト抽出 ---
def extract_data(text):
    patterns = {
        "製造番号": r"製造番号[:：]\s*([^\s)]+)",
        "印刷番号": r"印刷番号[:：]\s*([^\s\n]+)",
        "製造日": r"製造日[:：]\s*(.+?)(?:\n|$)",
        "会社名": r"会社名[:：]\s*(.+?)(?:\n|$)",
        "製品名": r"製品名[:：]\s*(.+?)(?:\n|$)",
        "製品種類": r"製品種類[:：]\s*(.+?)(?:\n|$)",
        "外装包材": r"外装包材[:：]\s*(.+?)(?:\n|$)",
        "表面印刷": r"表面印刷[:：][^\n]+.*?表面印刷[:：]\s*(.+?)(?:\n|$)",
        "製造個数": r"製造個数[:：]\s*(.+?)(?:\n|$)",
        "ファイル名": r"<印刷用データ\(\.FMT\)>.*?ファイル名[:：]\s*(.+?)(?:\n|$)",
        "印刷データ（元）": r"印刷データ[:：]\s*(.+?)(?:\n|$)"
    }

    results = {}
    for key, pattern in patterns.items():
        match = re.search(pattern, text, re.DOTALL)
        if match:
            value = match.group(1).strip()
            if key == "製造日":
                value = re.sub(r"^\d{4}/", "", value)  # ← ここで年を削除
            results[key] = value

    # 印刷データの新規・リピート分類
    if "印刷データ（元）" in results:
        raw = results.pop("印刷データ（元）")
        results["印刷データ"] = "リピート" if "従来の" in raw else "新規"
    else:
        results["印刷データ"] = ""

    # 前処理：改行なしで続くパターンを強制改行に変換
    text = re.sub(r"メモ欄[:：]\s*(会社共通情報[:：]|原料豆納品日[:：])", r"メモ欄：\n\1", text)

    # 抽出処理
    memo_match = re.search(r"メモ欄[:：]\s*\n([\s\S]*?)(?=\n?\s*(会社共通情報[:：]|原料豆納品日[:：]|$))", text)
    if memo_match:
        memo_text = memo_match.group(1).strip()
        results["メモ"] = "" if not memo_text or re.fullmatch(r"\s*", memo_text) else memo_text
    else:
        results["メモ"] = ""

    return results

# --- メインエンドポイント ---
@app.route("/", methods=["GET", "POST"])
def index():
    extracted_data = {}
    if request.method == "POST":
        text = request.form["text"]
        extracted_data = extract_data(text)
        logging.info("抽出データ: %s", extracted_data)

        # ✅ 1. GASでテンプレートブロック追加を実行
        try:
            response = requests.post(GAS_ENDPOINT, data={"mode": "copy"})
            logging.info("GAS Response: %s", response.text)
            response.raise_for_status()
            result = response.json()
            if result.get("status") == "OK":
                template_no = int(result.get("templateNumber"))
                # ✅ 追加部分
                block_index = template_no - 1
                # テンプレートデータ開始行の位置（A3 = 3行目）
                start_row = 3 + TEMPLATE_ROW_HEIGHT * block_index
                logging.info("テンプレNo: %d → 書き込み開始行: %d", template_no, start_row)
            else:
                flash("テンプレートの追加に失敗しました（GASから異常な応答）")
                return redirect(url_for("index"))
        except Exception as e:
            logging.error("GASテンプレート追加失敗: %s", str(e))
            flash("テンプレート追加時にエラー: " + str(e))
            return redirect(url_for("index"))

        # ✅ 2. Google Sheets に書き込み
        creds = get_credentials()
        client = gspread.authorize(creds)
        ss = client.open_by_key(SPREADSHEET_ID)
        output_ws = ss.worksheet("printlist")

        sheet_map = {
            "B4": "印刷データ",
            "C3": "ファイル名",
            "D3": "製造番号",
            "D7": "印刷番号",
            "E3": "製造日",
            "F3": "会社名",
            "F5": "製品名",
            "F8": "メモ",
            "H3": "製品種類",
            "H5": "外装包材",
            "H8": "表面印刷",
            "M3": "製造個数"
        }

        for cell_a1, key in sheet_map.items():
            if key in extracted_data:
                row = int(cell_a1[1:])
                col = ord(cell_a1[0].upper()) - 65 + 1
                target_row = start_row + (row - TEMPLATE_DATA_ROW_OFFSET)
                output_ws.update_cell(target_row, col, extracted_data[key])

        # ✅ 3. Excel ファイル書き出し（従来通り）
        wb = load_workbook(TEMPLATE_EXCEL_PATH)
        ws = wb.active

        cell_map = {
            "製造日": "B2",
            "製造番号": "E1",
            "印刷番号": "E2",
            "会社名": "B3",
            "製品名": "B4"
        }

        for key, cell in cell_map.items():
            if key in extracted_data:
                target_cell = ws[cell]
                if hasattr(target_cell, "merged_cells") or type(target_cell).__name__ == "MergedCell":
                    for merged_range in ws.merged_cells.ranges:
                        if cell in merged_range:
                            min_col, min_row, _, _ = range_boundaries(str(merged_range))
                            top_left_cell = ws.cell(row=min_row, column=min_col)
                            top_left_cell.value = extracted_data[key]
                            break
                else:
                    ws[cell] = extracted_data[key]

        excel_stream = io.BytesIO()
        wb.save(excel_stream)
        excel_stream.seek(0)

        return send_file(
            excel_stream,
            as_attachment=True,
            download_name="output.xlsx",
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    return render_template("index.html")

# --- GAS経由のスプレッドシートクリア ---
@app.route("/clear", methods=["POST"])
def clear_sheet():
    try:
        response = requests.post(GAS_ENDPOINT, data={"mode": "clear"})
        if response.status_code == 200 and response.text.strip() == "CLEARED":
            flash("スプレッドシートのデータをクリアしました。")
        else:
            flash("クリアに失敗しました：" + response.text)
    except Exception as e:
        flash("通信エラー：" + str(e))
    return redirect(url_for("index"))

# --- GAS経由のテンプレートブロックコピー ---
@app.route("/copy", methods=["POST"])
def copy_template_block():
    try:
        response = requests.post(GAS_ENDPOINT, data={"mode": "copy"})
        if response.status_code == 200:
            result = response.json()
            if result.get("status") == "OK":
                flash(f"テンプレートをコピーしました（No.{result.get('templateNumber')}）")
            else:
                flash("コピーに失敗しました（レスポンス異常）：" + str(result))
        else:
            flash("コピーに失敗しました：" + response.text)
    except Exception as e:
        flash("通信エラー：" + str(e))
    return redirect(url_for("index"))

if __name__ == "__main__":
    app.run(debug=True)
